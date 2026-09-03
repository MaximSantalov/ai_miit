#!/usr/bin/env python3
"""Генератор синтетического датасета вагонных перевозок для курса.

Детерминированно (фиксированный seed) создаёт:
  - data/logistics.db   — SQLite по схеме из internal/domain-conventions.md
  - data/dirty_orders.xlsx — намеренно «грязный» Excel для практики 4

Запуск:
    python data/generate_dataset.py

Скрипт идемпотентен: оба файла каждый раз пересоздаются с нуля.
Все данные полностью синтетические, реальные данные перевозчиков не используются.
"""

from __future__ import annotations

import datetime as dt
import math
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# --------------------------------------------------------------------------
# Константы и справочники (см. internal/domain-conventions.md)
# --------------------------------------------------------------------------

SEED = 42
RNG = np.random.default_rng(SEED)

ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "logistics.db"
XLSX_PATH = ROOT / "dirty_orders.xlsx"

HORIZON_START = dt.date(2025, 1, 1)
HORIZON_END = dt.date(2025, 12, 31)

N_STATIONS_TOTAL = 120
N_WAGONS = 3000
N_ORDERS = 800
TARGET_TRIPS = 12000

ROADS = [
    "Октябрьская",
    "Московская",
    "Горьковская",
    "Северная",
    "Северо-Кавказская",
    "Юго-Восточная",
    "Приволжская",
    "Куйбышевская",
    "Свердловская",
    "Южно-Уральская",
    "Западно-Сибирская",
    "Красноярская",
    "Восточно-Сибирская",
    "Забайкальская",
    "Дальневосточная",
]

ROAD_REGION = {
    "Октябрьская": "Северо-Запад",
    "Московская": "Центр",
    "Горьковская": "Приволжье (север)",
    "Северная": "Север",
    "Северо-Кавказская": "Юг",
    "Юго-Восточная": "Черноземье",
    "Приволжская": "Нижнее Поволжье",
    "Куйбышевская": "Среднее Поволжье",
    "Свердловская": "Урал",
    "Южно-Уральская": "Южный Урал",
    "Западно-Сибирская": "Кузбасс / Зап. Сибирь",
    "Красноярская": "Красноярский край",
    "Восточно-Сибирская": "Восточная Сибирь",
    "Забайкальская": "Забайкалье",
    "Дальневосточная": "Дальний Восток",
}

# Приблизительные центры дорог (для генерации координат станций).
ROAD_CENTER = {
    "Октябрьская": (59.9, 30.3),
    "Московская": (55.75, 37.6),
    "Горьковская": (56.3, 44.0),
    "Северная": (58.5, 39.8),
    "Северо-Кавказская": (47.2, 39.7),
    "Юго-Восточная": (51.7, 39.2),
    "Приволжская": (51.5, 46.0),
    "Куйбышевская": (53.2, 50.1),
    "Свердловская": (56.8, 60.6),
    "Южно-Уральская": (55.2, 61.4),
    "Западно-Сибирская": (55.0, 82.9),
    "Красноярская": (56.0, 92.9),
    "Восточно-Сибирская": (52.3, 104.3),
    "Забайкальская": (52.0, 113.5),
    "Дальневосточная": (48.5, 135.1),
}

WAGON_TYPES = [
    "полувагон",
    "крытый",
    "платформа",
    "цистерна",
    "хоппер-минераловоз",
    "фитинговая платформа",
    "думпкар",
]

CARGOS = [
    "уголь каменный",
    "кокс",
    "руда железная",
    "щебень",
    "зерно",
    "удобрения минеральные",
    "лес круглый",
    "металл в рулонах",
    "контейнеры",
    "нефтепродукты",
]

# Опорные станции из domain-conventions.md — обязаны присутствовать в stations.
# Формат: имя -> дорога.
OPORNYE_STATIONS = {
    "Кемерово-Сортировочное": "Западно-Сибирская",
    "Новокузнецк-Восточный": "Западно-Сибирская",
    "Инская": "Западно-Сибирская",
    "Алтайская": "Западно-Сибирская",
    "Мариинск": "Западно-Сибирская",
    "Тайшет": "Восточно-Сибирская",
    "Ачинск": "Красноярская",
    "Красноярск-Восточный": "Красноярская",
    "Иркутск-Сортировочный": "Восточно-Сибирская",
    "Улан-Удэ": "Восточно-Сибирская",
    "Чита-1": "Забайкальская",
    "Хабаровск-2": "Дальневосточная",
    "Находка-Восточная": "Дальневосточная",
    "Ванино": "Дальневосточная",
    "Новороссийск-Экспорт": "Северо-Кавказская",
    "Туапсе-Сортировочная": "Северо-Кавказская",
    "Ростов-Товарный": "Северо-Кавказская",
    "Лихая": "Северо-Кавказская",
    "Стойленская": "Юго-Восточная",
    "Старый Оскол": "Юго-Восточная",
    "Череповец-2": "Северная",
    "Костомукша": "Октябрьская",
    "Лужская": "Октябрьская",
    "Мурманск": "Октябрьская",
    "Екатеринбург-Сортировочный": "Свердловская",
    "Челябинск-Главный": "Южно-Уральская",
    "Магнитогорск-Грузовой": "Южно-Уральская",
    "Оренбург": "Южно-Уральская",
    "Саратов-Товарный": "Приволжская",
    "Волгоград-2": "Приволжская",
    "Сызрань-1": "Куйбышевская",
    "Уфа": "Куйбышевская",
    "Пермь-Сортировочная": "Свердловская",
    "Лоста": "Северная",
    "Ярославль-Главный": "Северная",
    "Орехово-Зуево": "Московская",
    "Бекасово-Сортировочное": "Московская",
    "Рыбное": "Московская",
}

# 12 опорных станций для полной матрицы Кузбасс <-> порты (нужна в практике 6, VRP).
# Координаты заданы вручную (приближены к реальной географии) — важно для
# правдоподобных дальних расстояний в задаче маршрутизации.
KUZBASS_PORT_12 = {
    "Кемерово-Сортировочное": (55.35, 86.05),
    "Новокузнецк-Восточный": (53.75, 87.15),
    "Инская": (55.05, 82.85),
    "Алтайская": (53.30, 83.55),
    "Мариинск": (56.21, 87.75),
    "Тайшет": (56.18, 97.97),
    "Ачинск": (56.28, 90.50),
    "Красноярск-Восточный": (56.05, 93.10),
    "Находка-Восточная": (42.83, 132.86),
    "Ванино": (49.09, 140.27),
    "Новороссийск-Экспорт": (44.72, 37.77),
    "Туапсе-Сортировочная": (44.10, 39.08),
}

KUZBASS_ORIGINS = [
    "Кемерово-Сортировочное",
    "Новокузнецк-Восточный",
    "Инская",
    "Алтайская",
    "Мариинск",
]
PORT_DESTINATIONS = [
    "Находка-Восточная",
    "Ванино",
    "Новороссийск-Экспорт",
    "Туапсе-Сортировочная",
]

CARGO_WAGON_TYPES = {
    "уголь каменный": ["полувагон"],
    "кокс": ["полувагон"],
    "руда железная": ["полувагон"],
    "щебень": ["полувагон", "думпкар"],
    "зерно": ["хоппер-минераловоз", "крытый"],
    "удобрения минеральные": ["хоппер-минераловоз", "крытый"],
    "лес круглый": ["платформа"],
    "металл в рулонах": ["платформа", "крытый"],
    "контейнеры": ["фитинговая платформа"],
    "нефтепродукты": ["цистерна"],
}

CAPACITY_BASE = {
    "полувагон": 70.0,
    "крытый": 68.0,
    "платформа": 65.0,
    "цистерна": 60.0,
    "хоппер-минераловоз": 70.0,
    "фитинговая платформа": 71.0,
    "думпкар": 60.0,
}

WAGON_TYPE_WEIGHTS = {
    "полувагон": 0.45,
    "цистерна": 0.15,
    "хоппер-минераловоз": 0.10,
    "платформа": 0.10,
    "фитинговая платформа": 0.08,
    "крытый": 0.08,
    "думпкар": 0.04,
}

OWNERS = [
    "СибТранс-Вагон",
    "ВостокЖелезТранс",
    "ЮжУралВагон",
    "ПрайдТранс",
    "МагистральЛизинг",
    "ТрансРесурс",
    "УралВагонСервис",
    "ДВ-ТрансОператор",
]

CLIENT_POOLS = {
    "уголь каменный": ["УгольТрансСервис", "Кузбасс-Экспорт", "СибУголь Логистика", "Разрез-Трейд"],
    "кокс": ["КоксХимТранс", "МеталлКокс Логистика"],
    "руда железная": ["РудаИнвест", "ГОК-Ресурс"],
    "щебень": ["НеруднаяКомпания", "КарьерТранс"],
    "зерно": ["АгроЗернопродукт", "ЗернoТрейд Юг"],
    "удобрения минеральные": ["ХимАгроСнаб", "АгроХимТранс"],
    "лес круглый": ["ЛесПром Сибирь", "ТаёжникЭкспорт"],
    "металл в рулонах": ["МеталлТрейд", "СтальХолдинг Логистика"],
    "контейнеры": ["КонтейнерЛайн", "ТрансКонт Сервис"],
    "нефтепродукты": ["НефтеТрансКомпани", "ПетроЛогистика"],
}

STATUS_WEIGHTS = {
    "закрыта": 0.55,
    "выполняется": 0.15,
    "подтверждена": 0.15,
    "новая": 0.10,
    "отменена": 0.05,
}

TYPE_SPEED_FACTOR = {
    "полувагон": 1.00,
    "цистерна": 1.08,
    "хоппер-минераловоз": 0.98,
    "платформа": 1.02,
    "фитинговая платформа": 1.10,
    "крытый": 0.92,
    "думпкар": 0.88,
}

CARGO_SPEED_FACTOR = {
    "уголь каменный": 0.95,
    "кокс": 0.97,
    "руда железная": 0.93,
    "щебень": 0.96,
    "зерно": 0.90,
    "удобрения минеральные": 0.94,
    "лес круглый": 0.97,
    "металл в рулонах": 1.00,
    "контейнеры": 1.12,
    "нефтепродукты": 1.06,
}

BASE_SPEED = 380.0  # км/сут — базовая скорость доставки (см. требования: 300-450 км/сут)
WINTER_FACTOR = 0.82
SUMMER_FACTOR = 1.08


# --------------------------------------------------------------------------
# Геометрия
# --------------------------------------------------------------------------

def haversine_km(lat1, lon1, lat2, lon2) -> float:
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def rail_distance_km(lat1, lon1, lat2, lon2, rng: np.random.Generator) -> float:
    straight = haversine_km(lat1, lon1, lat2, lon2)
    tortuosity = 1.28 + rng.normal(0, 0.03)
    km = straight * max(1.05, tortuosity)
    return max(60.0, round(km, 1))


# --------------------------------------------------------------------------
# 1. Дороги и станции
# --------------------------------------------------------------------------

def build_roads() -> pd.DataFrame:
    rows = []
    for i, name in enumerate(ROADS, start=1):
        rows.append({"road_id": i, "road_name": name, "region": ROAD_REGION[name]})
    return pd.DataFrame(rows)


NOUN_STEMS = [
    "Родник", "Ручей", "Курган", "Затон", "Бор", "Ельник", "Сосенки", "Полянка",
    "Взгорье", "Раздолье", "Плёс", "Яр", "Брод", "Луговая", "Склон", "Мыс",
    "Овражки", "Затишье", "Починок", "Выселки", "Берёзовка", "Дубровка",
    "Ольховка", "Сосновка", "Кедровка", "Тополёвка", "Ивановка", "Никольское",
    "Покровка", "Заводское", "Рудничная", "Каменка", "Песчаная", "Гранитная",
    "Известковая", "Медное", "Соляная", "Нефтяная", "Хлебная", "Заречное",
]
NAME_SUFFIXES = [
    "-Сортировочная", "-Товарная", "-Грузовая", "-Восточная", "-Западная",
    "-Южная", "-Северная", "-1", "-2", "-Промышленная",
]


def build_stations(roads_df: pd.DataFrame, rng: np.random.Generator) -> tuple[pd.DataFrame, dict]:
    road_id_by_name = dict(zip(roads_df["road_name"], roads_df["road_id"]))
    used_names = set(OPORNYE_STATIONS.keys())
    rows = []
    code_counter = {r: 1000 for r in ROADS}
    name_to_code: dict[str, str] = {}

    def next_code(road_name: str) -> str:
        road_id = road_id_by_name[road_name]
        code_counter[road_name] += 1
        return f"{road_id:02d}{code_counter[road_name]:04d}"

    def coords_for(road_name: str, exact: tuple[float, float] | None) -> tuple[float, float]:
        if exact is not None:
            return exact
        clat, clon = ROAD_CENTER[road_name]
        lat = clat + rng.normal(0, 1.3)
        lon = clon + rng.normal(0, 2.2)
        return round(lat, 4), round(lon, 4)

    # опорные станции (обязательны все)
    for name, road_name in OPORNYE_STATIONS.items():
        exact = KUZBASS_PORT_12.get(name)
        lat, lon = coords_for(road_name, exact)
        code = next_code(road_name)
        name_to_code[name] = code
        rows.append({
            "station_code": code, "station_name": name, "road_id": road_id_by_name[road_name],
            "lat": lat, "lon": lon,
        })

    # синтетические станции, чтобы добрать до N_STATIONS_TOTAL, равномерно по дорогам
    n_extra = N_STATIONS_TOTAL - len(rows)
    per_road = max(1, n_extra // len(ROADS))
    counts = {r: per_road for r in ROADS}
    remainder = n_extra - per_road * len(ROADS)
    for r in ROADS[:remainder]:
        counts[r] += 1

    for road_name in ROADS:
        for _ in range(counts[road_name]):
            for _try in range(200):
                noun = rng.choice(NOUN_STEMS)
                suffix = rng.choice(NAME_SUFFIXES)
                candidate = f"{noun}{suffix}"
                if candidate not in used_names:
                    used_names.add(candidate)
                    break
            lat, lon = coords_for(road_name, None)
            code = next_code(road_name)
            name_to_code[candidate] = code
            rows.append({
                "station_code": code, "station_name": candidate, "road_id": road_id_by_name[road_name],
                "lat": lat, "lon": lon,
            })

    df = pd.DataFrame(rows)
    return df, name_to_code


# --------------------------------------------------------------------------
# 2. Вагоны
# --------------------------------------------------------------------------

def build_wagons(rng: np.random.Generator) -> pd.DataFrame:
    types = list(WAGON_TYPE_WEIGHTS.keys())
    weights = np.array([WAGON_TYPE_WEIGHTS[t] for t in types])
    weights = weights / weights.sum()

    wagon_types = rng.choice(types, size=N_WAGONS, p=weights)

    numbers = set()
    wagon_nos = []
    while len(wagon_nos) < N_WAGONS:
        n = int(rng.integers(20_000_000, 89_999_999))
        s = f"{n:08d}"
        if s not in numbers:
            numbers.add(s)
            wagon_nos.append(s)

    rows = []
    today = HORIZON_END + dt.timedelta(days=60)  # ~2026-03-01
    for i in range(N_WAGONS):
        wtype = wagon_types[i]
        capacity = round(CAPACITY_BASE[wtype] + rng.normal(0, 1.5), 1)
        owner = rng.choice(OWNERS)
        build_year = int(round(rng.triangular(1998, 2020, 2024)))
        build_date = dt.date(build_year, int(rng.integers(1, 13)), int(rng.integers(1, 28)))
        earliest_repair = max(build_date, dt.date(2022, 1, 1))
        span = (today - earliest_repair).days
        span = max(span, 1)
        last_repair = earliest_repair + dt.timedelta(days=int(rng.integers(0, span)))
        rows.append({
            "wagon_no": wagon_nos[i],
            "wagon_type": wtype,
            "capacity_t": capacity,
            "owner": owner,
            "build_year": build_year,
            "last_repair_date": last_repair.isoformat(),
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 3. Заявки (orders)
# --------------------------------------------------------------------------

def random_date_in_horizon(rng: np.random.Generator, start: dt.date = HORIZON_START, end: dt.date = HORIZON_END) -> dt.date:
    span = (end - start).days
    return start + dt.timedelta(days=int(rng.integers(0, span + 1)))


def build_orders(stations_df: pd.DataFrame, name_to_code: dict, rng: np.random.Generator) -> pd.DataFrame:
    all_names = stations_df["station_name"].tolist()
    cargo_list = list(CARGO_WAGON_TYPES.keys())
    statuses = list(STATUS_WEIGHTS.keys())
    status_p = np.array([STATUS_WEIGHTS[s] for s in statuses])
    status_p = status_p / status_p.sum()

    rows = []
    for i in range(1, N_ORDERS + 1):
        is_coal_flow = rng.random() < 0.30
        if is_coal_flow:
            cargo = "уголь каменный"
            wagon_type = "полувагон"
            from_name = rng.choice(KUZBASS_ORIGINS)
            to_name = rng.choice(PORT_DESTINATIONS)
            wagons_needed = int(rng.integers(10, 41))
        else:
            cargo = rng.choice(cargo_list)
            wagon_type = rng.choice(CARGO_WAGON_TYPES[cargo])
            from_name, to_name = rng.choice(all_names, size=2, replace=False)
            wagons_needed = int(rng.integers(3, 13))

        client = rng.choice(CLIENT_POOLS[cargo])
        date_from = random_date_in_horizon(rng, HORIZON_START, dt.date(2025, 12, 15))
        window = int(rng.integers(3, 16))
        date_to = min(date_from + dt.timedelta(days=window), HORIZON_END)
        lead = int(rng.integers(1, 21))
        created_at = max(date_from - dt.timedelta(days=lead), HORIZON_START)
        status = rng.choice(statuses, p=status_p)

        rows.append({
            "order_id": i,
            "client": client,
            "cargo": cargo,
            "wagon_type": wagon_type,
            "wagons_needed": wagons_needed,
            "from_station": name_to_code[from_name],
            "to_station": name_to_code[to_name],
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "created_at": created_at.isoformat(),
            "status": status,
            "_from_name": from_name,
            "_to_name": to_name,
        })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 4. Рейсы (trips) + distances
# --------------------------------------------------------------------------

def duration_plan(distance_km: float, rng: np.random.Generator) -> float:
    dwell = 0.7
    noise = rng.normal(0, 0.25)
    return max(0.5, distance_km / BASE_SPEED + dwell + noise)


def duration_fact(distance_km: float, wagon_type: str, cargo: str | None, month: int, rng: np.random.Generator) -> float:
    type_factor = TYPE_SPEED_FACTOR[wagon_type]
    cargo_factor = CARGO_SPEED_FACTOR.get(cargo, 1.0)
    season_factor = WINTER_FACTOR if month in (12, 1, 2) else (SUMMER_FACTOR if month in (6, 7, 8) else 1.0)
    effective_speed = BASE_SPEED * type_factor * cargo_factor * season_factor
    dwell = max(0.1, rng.normal(1.0, 0.9))
    base = distance_km / effective_speed + dwell
    noise = rng.normal(0, 2.3)
    duration = base + noise
    if rng.random() < 0.15:
        duration += rng.exponential(5.0)
    return max(0.3, duration)


def build_trips_and_distances(stations_df: pd.DataFrame, orders_df: pd.DataFrame, wagons_df: pd.DataFrame,
                               name_to_code: dict, rng: np.random.Generator):
    coord = {row.station_code: (row.lat, row.lon) for row in stations_df.itertuples()}
    dist_cache: dict[tuple[str, str], float] = {}

    def distance_between(code_a: str, code_b: str) -> float:
        key = (code_a, code_b)
        if key in dist_cache:
            return dist_cache[key]
        lat1, lon1 = coord[code_a]
        lat2, lon2 = coord[code_b]
        km = rail_distance_km(lat1, lon1, lat2, lon2, rng)
        dist_cache[key] = km
        dist_cache[(code_b, code_a)] = km
        return km

    wagons_by_type: dict[str, list[str]] = {}
    for row in wagons_df.itertuples():
        wagons_by_type.setdefault(row.wagon_type, []).append(row.wagon_no)

    trip_rows = []
    trip_id = 1

    for order in orders_df.itertuples():
        n_trips = min(order.wagons_needed, 45)
        pool = wagons_by_type.get(order.wagon_type, wagons_df["wagon_no"].tolist())
        date_from = dt.date.fromisoformat(order.date_from)
        date_to = dt.date.fromisoformat(order.date_to)
        span = max((date_to - date_from).days, 0)
        dist_km = distance_between(order.from_station, order.to_station)

        for _ in range(n_trips):
            wagon_no = pool[int(rng.integers(0, len(pool)))]
            depart = date_from + dt.timedelta(days=int(rng.integers(0, span + 1)))
            month = depart.month
            plan_days = duration_plan(dist_km, rng)
            fact_days = duration_fact(dist_km, order.wagon_type, order.cargo, month, rng)
            arrive_plan = depart + dt.timedelta(days=round(plan_days))
            arrive_fact_date = depart + dt.timedelta(days=round(fact_days))
            arrive_fact = arrive_fact_date.isoformat() if arrive_fact_date <= HORIZON_END else None

            trip_rows.append({
                "trip_id": trip_id,
                "wagon_no": wagon_no,
                "order_id": order.order_id,
                "from_station": order.from_station,
                "to_station": order.to_station,
                "depart_dt": depart.isoformat(),
                "arrive_dt_plan": arrive_plan.isoformat(),
                "arrive_dt_fact": arrive_fact,
                "distance_km": dist_km,
            })
            trip_id += 1

    order_trip_count = len(trip_rows)
    reposition_target = max(0, TARGET_TRIPS - order_trip_count)

    all_names = stations_df["station_name"].tolist()
    all_codes_by_name = name_to_code
    for _ in range(reposition_target):
        if rng.random() < 0.6:
            # порожний возврат по доминирующему угольному направлению порт -> Кузбасс
            from_name = rng.choice(PORT_DESTINATIONS)
            to_name = rng.choice(KUZBASS_ORIGINS)
            wagon_type = "полувагон"
        else:
            from_name, to_name = rng.choice(all_names, size=2, replace=False)
            types = list(WAGON_TYPE_WEIGHTS.keys())
            weights = np.array([WAGON_TYPE_WEIGHTS[t] for t in types])
            weights = weights / weights.sum()
            wagon_type = rng.choice(types, p=weights)

        from_code = all_codes_by_name[from_name]
        to_code = all_codes_by_name[to_name]
        pool = wagons_by_type.get(wagon_type, wagons_df["wagon_no"].tolist())
        wagon_no = pool[int(rng.integers(0, len(pool)))]
        depart = random_date_in_horizon(rng)
        month = depart.month
        dist_km = distance_between(from_code, to_code)
        plan_days = duration_plan(dist_km, rng)
        fact_days = duration_fact(dist_km, wagon_type, None, month, rng)
        arrive_plan = depart + dt.timedelta(days=round(plan_days))
        arrive_fact_date = depart + dt.timedelta(days=round(fact_days))
        arrive_fact = arrive_fact_date.isoformat() if arrive_fact_date <= HORIZON_END else None

        trip_rows.append({
            "trip_id": trip_id,
            "wagon_no": wagon_no,
            "order_id": None,
            "from_station": from_code,
            "to_station": to_code,
            "depart_dt": depart.isoformat(),
            "arrive_dt_plan": arrive_plan.isoformat(),
            "arrive_dt_fact": arrive_fact,
            "distance_km": dist_km,
        })
        trip_id += 1

    trips_df = pd.DataFrame(trip_rows)

    # distances: все пары, реально встретившиеся в trips, + полная матрица 12 опорных станций
    pairs = set(zip(trips_df["from_station"], trips_df["to_station"]))
    for name_a in KUZBASS_PORT_12:
        for name_b in KUZBASS_PORT_12:
            if name_a == name_b:
                continue
            pairs.add((name_to_code[name_a], name_to_code[name_b]))

    dist_rows = []
    for a, b in sorted(pairs):
        dist_rows.append({"from_station": a, "to_station": b, "km": distance_between(a, b)})
    distances_df = pd.DataFrame(dist_rows)

    return trips_df, distances_df


# --------------------------------------------------------------------------
# 5. Дислокация (снимки операций по вагонам)
# --------------------------------------------------------------------------

OPERATIONS = ["погрузка", "выгрузка", "прибытие", "отправление", "простой"]


def build_dislocation(stations_df: pd.DataFrame, wagons_df: pd.DataFrame, rng: np.random.Generator) -> pd.DataFrame:
    codes = stations_df["station_code"].tolist()
    rows = []
    rec_id = 1
    for w in wagons_df.itertuples():
        n_records = int(rng.integers(2, 5))
        for _ in range(n_records):
            op = rng.choice(OPERATIONS)
            station_code = codes[int(rng.integers(0, len(codes)))]
            op_date = random_date_in_horizon(rng)
            if op in ("погрузка", "отправление"):
                is_loaded = 1
                cargo = rng.choice(CARGOS)
            elif op == "выгрузка":
                is_loaded = 0
                cargo = None
            else:
                is_loaded = int(rng.integers(0, 2))
                cargo = rng.choice(CARGOS) if is_loaded else None
            dest_code = None
            if op == "отправление":
                dest_code = codes[int(rng.integers(0, len(codes)))]
            rows.append({
                "id": rec_id,
                "wagon_no": w.wagon_no,
                "station_code": station_code,
                "operation": op,
                "operation_dt": op_date.isoformat(),
                "is_loaded": is_loaded,
                "cargo": cargo,
                "dest_station_code": dest_code,
            })
            rec_id += 1
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------
# 6. Запись в SQLite
# --------------------------------------------------------------------------

def write_sqlite(roads_df, stations_df, wagons_df, orders_df, trips_df, distances_df, dislocation_df):
    if DB_PATH.exists():
        DB_PATH.unlink()
    conn = sqlite3.connect(DB_PATH)
    try:
        orders_out = orders_df.drop(columns=["_from_name", "_to_name"])
        roads_df.to_sql("roads", conn, index=False)
        stations_df.to_sql("stations", conn, index=False)
        wagons_df.to_sql("wagons", conn, index=False)
        dislocation_df.to_sql("dislocation", conn, index=False)
        orders_out.to_sql("orders", conn, index=False)
        trips_df.to_sql("trips", conn, index=False)
        distances_df.to_sql("distances", conn, index=False)

        cur = conn.cursor()
        cur.execute("CREATE INDEX idx_stations_road ON stations(road_id)")
        cur.execute("CREATE INDEX idx_trips_order ON trips(order_id)")
        cur.execute("CREATE INDEX idx_trips_wagon ON trips(wagon_no)")
        cur.execute("CREATE INDEX idx_dislocation_wagon ON dislocation(wagon_no)")
        cur.execute("CREATE INDEX idx_distances_pair ON distances(from_station, to_station)")
        conn.commit()
    finally:
        conn.close()


# --------------------------------------------------------------------------
# 7. dirty_orders.xlsx для практики 4
# --------------------------------------------------------------------------

def _fmt_date(d: dt.date, style: int) -> str:
    if style == 0:
        return d.strftime("%d.%m.%Y")
    if style == 1:
        return d.strftime("%Y-%m-%d")
    # style 2 — текстовое приближение с годом, style 3 — то же без года
    months = {
        1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
        7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря",
    }
    decade = "начало" if d.day <= 10 else ("середина" if d.day <= 20 else "конец")
    if style == 3:
        # Текстовая дата БЕЗ года. Ключевой случай практики 4: год восстанавливается
        # по правилу («не раньше даты формирования документа»), а не додумывается моделью.
        return f"{decade} {months[d.month]}"
    return f"{decade} {months[d.month]} {d.year}"


def _fmt_wagons(n: int, style: int) -> str:
    if style == 0:
        return str(n)
    if style == 1:
        return f"{n} шт"
    if style == 2:
        return f"{n} ваг."
    return ""  # пусто


def _typo(name: str) -> str:
    # детерминированная типовая опечатка: перестановка двух соседних букв около середины слова
    if len(name) < 3:
        return name
    idx = len(name) // 2
    chars = list(name)
    for offset in range(len(chars) - 1):
        i = idx + (offset // 2 if offset % 2 == 0 else -(offset // 2 + 1))
        if 0 <= i < len(chars) - 1 and chars[i] != chars[i + 1] and chars[i].isalpha() and chars[i + 1].isalpha():
            chars[i], chars[i + 1] = chars[i + 1], chars[i]
            return "".join(chars)
    return name + name[-1]


def build_dirty_orders(stations_df: pd.DataFrame, rng: np.random.Generator) -> list[str]:
    """Строит data/dirty_orders.xlsx. Возвращает журнал дефектов (для README)."""
    all_names = stations_df["station_name"].tolist()
    cargo_list = list(CARGO_WAGON_TYPES.keys())
    defects_log: list[str] = []

    def make_records(n: int, cargo_bias_coal: float) -> list[dict]:
        recs = []
        for _ in range(n):
            if rng.random() < cargo_bias_coal:
                cargo = "уголь каменный"
                wagon_type = "полувагон"
                from_name = rng.choice(KUZBASS_ORIGINS)
                to_name = rng.choice(PORT_DESTINATIONS)
            else:
                cargo = rng.choice(cargo_list)
                wagon_type = rng.choice(CARGO_WAGON_TYPES[cargo])
                from_name, to_name = rng.choice(all_names, size=2, replace=False)
            client = rng.choice(CLIENT_POOLS[cargo])
            wagons = int(rng.integers(3, 60))
            date_from = random_date_in_horizon(rng)
            recs.append({
                "client": client, "cargo": cargo, "wagon_type": wagon_type,
                "from_name": from_name, "to_name": to_name,
                "wagons": wagons, "date": date_from,
            })
        return recs

    wb = Workbook()

    # ---------------- Лист 1: «Реестр_заявок_1» ----------------
    ws1 = wb.active
    ws1.title = "Реестр_заявок_1"
    recs1 = make_records(45, cargo_bias_coal=0.35)

    ws1["A1"] = "ООО «ТрансЛогистика» — реестр заявок на подачу вагонов"
    ws1["A1"].font = Font(bold=True)
    ws1["A2"] = "Дата формирования отчёта: 01.03.2026"
    ws1["A3"] = "Утвердил: нач. отдела перевозок Соколов А.П."
    ws1["A4"] = None  # пустая строка-разделитель

    ws1.merge_cells("E5:F5")
    ws1["E5"] = "Маршрут"
    ws1["E5"].alignment = Alignment(horizontal="center")
    headers1 = ["№", "Клиент", "Груз", "Род вагона", "Ст. отправления", "Ст. назначения", "Кол-во вагонов", "Дата подачи"]
    for col, title in enumerate(headers1, start=1):
        ws1.cell(row=6, column=col, value=title).font = Font(bold=True)
    defects_log.append(
        "Лист 'Реестр_заявок_1': строки 1-4 — шапка документа (название/дата/подпись/пустая строка) "
        "перед таблицей; строка 5 — объединённые ячейки E5:F5 ('Маршрут') над строкой заголовков 6."
    )

    r = 7
    written_rows: list[tuple[int, dict]] = []
    dup_sources: list[tuple[int, dict]] = []
    for i, rec in enumerate(recs1):
        date_style = i % 4
        wagons_style = i % 4
        from_name = rec["from_name"]
        to_name = rec["to_name"]
        client = rec["client"]
        cargo = rec["cargo"]
        wagon_type = rec["wagon_type"]

        if i == 8:
            from_name = _typo(from_name)
        if i == 22:
            to_name = _typo(to_name)
        if i == 3:
            client = f"  {client}  "
        if i == 11:
            cargo = cargo.upper()

        if i == 15:
            r += 1  # пустая строка-разделитель внутри таблицы
            defects_log.append(f"Лист 'Реестр_заявок_1': строка {r} — пустая строка-разделитель внутри таблицы.")

        row_vals = [i + 1, client, cargo, wagon_type, from_name, to_name,
                    _fmt_wagons(rec["wagons"], wagons_style), _fmt_date(rec["date"], date_style)]
        for col, v in enumerate(row_vals, start=1):
            ws1.cell(row=r, column=col, value=v)
        written_rows.append((r, dict(rec, from_name=from_name, to_name=to_name, client=client,
                                      cargo=cargo, wagon_type=wagon_type,
                                      wagons_str=row_vals[6], date_str=row_vals[7])))
        if i in (2, 20):
            dup_sources.append((r, dict(zip(
                ["order_no", "client", "cargo", "wagon_type", "from_name", "to_name", "wagons_str", "date_str"],
                row_vals))))
        r += 1

    dup_rows_written = []
    for src_row, src_vals in dup_sources:
        for col, v in enumerate(list(src_vals.values()), start=1):
            ws1.cell(row=r, column=col, value=v)
        dup_rows_written.append((r, src_row))
        r += 1
    defects_log.append(
        "Лист 'Реестр_заявок_1': строки {} — точные дубли строк {} соответственно (полные дубли записей).".format(
            ", ".join(str(x[0]) for x in dup_rows_written), ", ".join(str(x[1]) for x in dup_rows_written)
        )
    )

    total_wagons_1 = sum(rec["wagons"] for rec in recs1)
    ws1.cell(row=r, column=1, value="ИТОГО")
    ws1.cell(row=r, column=7, value=total_wagons_1)
    defects_log.append(f"Лист 'Реестр_заявок_1': строка {r} — итоговая строка 'ИТОГО' внутри табличного диапазона.")
    defects_log.append(
        "Лист 'Реестр_заявок_1': даты в столбце H чередуются по четырём форматам "
        "(01.03.2025 / 2025-03-01 / 'начало марта 2025' / 'начало марта' без года) "
        "каждые 4 строки; "
        "количество вагонов в столбце G чередуется по 4 стилям ('15' / '15 шт' / '15 ваг.' / пусто)."
    )
    defects_log.append(
        f"Лист 'Реестр_заявок_1': строка {written_rows[8][0]} — опечатка в станции отправления "
        f"('{written_rows[8][1]['from_name']}'); строка {written_rows[22][0]} — опечатка в станции назначения "
        f"('{written_rows[22][1]['to_name']}')."
    )
    defects_log.append(
        f"Лист 'Реестр_заявок_1': строка {written_rows[3][0]} — лишние пробелы вокруг клиента; "
        f"строка {written_rows[11][0]} — груз в верхнем регистре."
    )

    # ---------------- Лист 2: «Реестр_заявок_2» ----------------
    ws2 = wb.create_sheet("Реестр_заявок_2")
    recs2 = make_records(40, cargo_bias_coal=0.3)
    ws2["A1"] = "Реестр заявок за март 2026 — филиал «Юг»"
    ws2["A2"] = None
    headers2 = ["Заказчик", "Груз", "Тип вагона", "Станция отпр.", "Станция назн.", "Вагонов, шт", "Срок подачи"]
    for col, title in enumerate(headers2, start=1):
        ws2.cell(row=3, column=col, value=title).font = Font(bold=True)
    defects_log.append(
        "Лист 'Реестр_заявок_2': строки 1-2 — укороченная шапка документа (название + пустая строка); "
        "строка 3 — заголовки с именами колонок, отличными от листа 1 "
        "('Заказчик' вместо 'Клиент', 'Станция отпр.' вместо 'Ст. отправления' и т. д.)."
    )

    r = 4
    written2 = []
    for i, rec in enumerate(recs2):
        date_style = (i + 1) % 4
        wagons_style = (i + 2) % 4
        from_name = rec["from_name"]
        to_name = rec["to_name"]
        cargo = rec["cargo"]
        if i == 5:
            from_name = _typo(from_name)
        if i == 18:
            cargo = f" {cargo}"
        if i == 25:
            r += 1
            defects_log.append(f"Лист 'Реестр_заявок_2': строка {r} — пустая строка-разделитель внутри таблицы.")
        row_vals = [rec["client"], cargo, rec["wagon_type"], from_name, to_name,
                    _fmt_wagons(rec["wagons"], wagons_style), _fmt_date(rec["date"], date_style)]
        for col, v in enumerate(row_vals, start=1):
            ws2.cell(row=r, column=col, value=v)
        written2.append((r, row_vals))
        r += 1

    # ещё 1-2 дубля на листе 2
    for src_idx in (4,):
        src_row, src_vals = written2[src_idx]
        for col, v in enumerate(src_vals, start=1):
            ws2.cell(row=r, column=col, value=v)
        defects_log.append(f"Лист 'Реестр_заявок_2': строка {r} — точный дубль строки {src_row}.")
        r += 1

    total_wagons_2 = sum(rec["wagons"] for rec in recs2)
    ws2.cell(row=r, column=1, value="ИТОГО")
    ws2.cell(row=r, column=6, value=total_wagons_2)
    defects_log.append(f"Лист 'Реестр_заявок_2': строка {r} — итоговая строка 'ИТОГО'.")
    defects_log.append(
        f"Лист 'Реестр_заявок_2': строка {written2[5][0]} — опечатка в станции отправления "
        f"('{written2[5][1][3]}'); строка {written2[18][0]} — ведущий пробел в названии груза."
    )

    # ---------------- Лист 3: «Выгрузка_raw» (без заголовков) ----------------
    ws3 = wb.create_sheet("Выгрузка_raw")
    recs3 = make_records(35, cargo_bias_coal=0.25)
    defects_log.append(
        "Лист 'Выгрузка_raw': заголовков нет вовсе, данные начинаются с первой строки. "
        "Фиксированный порядок колонок: A=Клиент, B=Груз, C=Род вагона, D=откуда (станция отправления), "
        "E=куда (станция назначения), F=Вагоны, G=Дата."
    )
    r = 1
    written3 = []
    for i, rec in enumerate(recs3):
        date_style = i % 4
        wagons_style = (i + 3) % 4
        from_name = rec["from_name"]
        if i == 12:
            from_name = _typo(from_name)
        row_vals = [rec["client"], rec["cargo"], rec["wagon_type"], from_name, rec["to_name"],
                    _fmt_wagons(rec["wagons"], wagons_style), _fmt_date(rec["date"], date_style)]
        for col, v in enumerate(row_vals, start=1):
            ws3.cell(row=r, column=col, value=v)
        written3.append((r, row_vals))
        r += 1
    # дубль на листе 3
    src_row, src_vals = written3[7]
    for col, v in enumerate(src_vals, start=1):
        ws3.cell(row=r, column=col, value=v)
    defects_log.append(f"Лист 'Выгрузка_raw': строка {r} — точный дубль строки {src_row}.")
    defects_log.append(f"Лист 'Выгрузка_raw': строка {written3[12][0]} — опечатка в станции отправления.")

    wb.save(XLSX_PATH)
    return defects_log


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def main():
    roads_df = build_roads()
    stations_df, name_to_code = build_stations(roads_df, RNG)
    wagons_df = build_wagons(RNG)
    orders_df = build_orders(stations_df, name_to_code, RNG)
    trips_df, distances_df = build_trips_and_distances(stations_df, orders_df, wagons_df, name_to_code, RNG)
    dislocation_df = build_dislocation(stations_df, wagons_df, RNG)

    write_sqlite(roads_df, stations_df, wagons_df, orders_df, trips_df, distances_df, dislocation_df)
    defects_log = build_dirty_orders(stations_df, np.random.default_rng(SEED + 1))

    scratch_defects = Path(__file__).resolve().parent.parent / ".scratch" / "dirty_orders_defects.txt"
    scratch_defects.parent.mkdir(exist_ok=True)
    scratch_defects.write_text("\n".join(defects_log), encoding="utf-8")

    print("=== Сводка по data/logistics.db ===")
    for name, df in [
        ("roads", roads_df), ("stations", stations_df), ("wagons", wagons_df),
        ("dislocation", dislocation_df), ("orders", orders_df.drop(columns=["_from_name", "_to_name"])),
        ("trips", trips_df), ("distances", distances_df),
    ]:
        print(f"  {name}: {len(df)} строк")
    print(f"=== data/dirty_orders.xlsx: 3 листа, всего дефектов задокументировано: {len(defects_log)} ===")
    print(f"DB path: {DB_PATH}")
    print(f"XLSX path: {XLSX_PATH}")


if __name__ == "__main__":
    main()
